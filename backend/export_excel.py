"""Excel 导出：openpyxl 生成 xlsx（不用 csv——中文 Excel 打开会乱码）"""
import io
import json
from datetime import date, timedelta
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from models import SalesOrder, PurchaseOrder, RawMaterial, Product, ProductionRecord, Customer, Supplier

HEADER_FILL = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def _new_sheet(wb: Workbook, title: str, headers: List[str], rows: List[list]):
    ws = wb.active if wb.active.title == "Sheet" and wb.active.max_row == 1 else wb.create_sheet(title)
    ws.title = title
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(row)
    for idx, header in enumerate(headers, 1):
        width = max(len(str(header)) * 2.2, *(len(str(r[idx - 1])) * 1.9 for r in rows)) if rows else len(str(header)) * 2.2
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 10), 50)
    ws.freeze_panes = "A2"
    return ws


def _to_xlsx_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_sales(db: Session, start_date: str = "", end_date: str = "") -> bytes:
    today = date.today()
    if not start_date:
        start_date = (today.replace(day=1)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = today.strftime("%Y-%m-%d")
    orders = db.query(SalesOrder).filter(
        SalesOrder.date >= start_date, SalesOrder.date <= end_date,
    ).order_by(SalesOrder.date.desc(), SalesOrder.id.desc()).all()

    wb = Workbook()
    rows = []
    for o in orders:
        cname = o.customer.name if o.customer else ""
        items = json.loads(o.items) if o.items else []
        items_text = "；".join(f"{i.get('product_name', '')}×{i.get('quantity', 0)}{i.get('unit', '')}" for i in items)
        rows.append([o.order_no or "", str(o.date), cname, items_text, o.total_amount or 0,
                     o.status, o.payment_status, o.paid_amount or 0,
                     round((o.total_amount or 0) - (o.paid_amount or 0), 2),
                     o.operator or "", o.notes or ""])
    _new_sheet(wb, "销售明细", ["订单号", "日期", "客户", "产品明细", "订单金额", "状态", "付款状态", "已付金额", "未付金额", "操作人", "备注"], rows)
    return _to_xlsx_bytes(wb)


def export_purchases(db: Session, start_date: str = "", end_date: str = "") -> bytes:
    today = date.today()
    if not start_date:
        start_date = (today - timedelta(days=89)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = today.strftime("%Y-%m-%d")
    orders = db.query(PurchaseOrder).filter(
        PurchaseOrder.date >= start_date, PurchaseOrder.date <= end_date,
    ).order_by(PurchaseOrder.date.desc(), PurchaseOrder.id.desc()).all()

    wb = Workbook()
    rows = []
    for o in orders:
        sname = o.supplier.name if o.supplier else ""
        items = json.loads(o.items) if o.items else []
        items_text = "；".join(f"{i.get('material_name', '')}×{i.get('quantity', 0)}{i.get('unit', '')}" for i in items)
        rows.append([o.order_no or "", str(o.date), sname, items_text, o.total_amount or 0,
                     o.status, o.operator or "", o.notes or ""])
    _new_sheet(wb, "采购明细", ["采购单号", "日期", "供应商", "原料明细", "金额", "状态", "操作人", "备注"], rows)
    return _to_xlsx_bytes(wb)


def export_inventory(db: Session) -> bytes:
    materials = db.query(RawMaterial).order_by(RawMaterial.category, RawMaterial.name).all()
    products = db.query(Product).order_by(Product.category, Product.name).all()

    wb = Workbook()
    mat_rows = [[m.name or "", m.category or "", m.current_stock or 0, m.unit or "",
                 m.safety_stock or 0, m.purchase_price or 0,
                 round((m.current_stock or 0) * (m.purchase_price or 0), 2),
                 m.supplier or ""] for m in materials]
    _new_sheet(wb, "原料库存", ["原料名称", "类别", "当前库存", "单位", "安全线", "采购价", "库存价值", "供应商"], mat_rows)

    prod_rows = [[p.name or "", p.category or "", p.spec or "", p.current_stock or 0, p.unit or ""] for p in products]
    _new_sheet(wb, "产品库存", ["产品名称", "类别", "规格", "当前库存", "单位"], prod_rows)
    return _to_xlsx_bytes(wb)


def export_receivables(db: Session) -> bytes:
    orders = db.query(SalesOrder).filter(
        SalesOrder.payment_status != "已付款",
        SalesOrder.status != "已取消",
    ).order_by(SalesOrder.date.asc()).all()

    wb = Workbook()
    rows = []
    today = date.today()
    for o in orders:
        cname = o.customer.name if o.customer else ""
        unpaid = round((o.total_amount or 0) - (o.paid_amount or 0), 2)
        overdue_days = (today - o.date).days
        rows.append([o.order_no or "", str(o.date), cname, o.total_amount or 0,
                     o.paid_amount or 0, unpaid, o.payment_status, o.status,
                     overdue_days if overdue_days > 30 else ""])
    _new_sheet(wb, "应收账款", ["订单号", "日期", "客户", "订单金额", "已付金额", "未付金额", "付款状态", "订单状态", "逾期天数(>30天)"], rows)
    return _to_xlsx_bytes(wb)


def export_production(db: Session, start_date: str = "", end_date: str = "") -> bytes:
    today = date.today()
    if not start_date:
        start_date = (today - timedelta(days=29)).strftime("%Y-%m-%d")
    if not end_date:
        end_date = today.strftime("%Y-%m-%d")
    records = db.query(ProductionRecord).filter(
        ProductionRecord.date >= start_date, ProductionRecord.date <= end_date,
    ).order_by(ProductionRecord.date.desc(), ProductionRecord.id.desc()).all()

    wb = Workbook()
    material_names = {m.id: m.name for m in db.query(RawMaterial.id, RawMaterial.name).all()}
    rows = []
    for r in records:
        pname = r.product.name if r.product else ""
        used = json.loads(r.raw_materials_used) if r.raw_materials_used else []
        used_text = "；".join(
            f"{material_names.get(u.get('material_id'), '')}×{u.get('quantity', 0)}" for u in used
        ) if used else ""
        rows.append([str(r.date), pname, r.quantity or 0, r.unit or "",
                     r.sugar_degree if r.sugar_degree is not None else "",
                     used_text, r.operator or "", r.notes or ""])
    _new_sheet(wb, "生产记录", ["日期", "产品", "产量", "单位", "糖度", "消耗原料", "操作人", "备注"], rows)
    return _to_xlsx_bytes(wb)
